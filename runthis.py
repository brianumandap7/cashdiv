from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog
import os

# Step 1: Let the user select the Excel file via a file dialog
root = tk.Tk()
root.withdraw()  # Hide the main tkinter window
excel_file_path = filedialog.askopenfilename(
    title="Select an Excel file",
    filetypes=[("Excel files", "*.xlsx;*.xls")]
)

# Ensure the user selected a file
if not excel_file_path:
    print("No file selected, exiting...")
    exit()

wb = load_workbook(excel_file_path)
sheet = wb.active  # Or specify the sheet name: wb['Sheet1']

replacement_string = '9999999999DEPARTMENT OF TRANSPORTAT000001079200266000010633854605905300000 00000'
blank_spaces = "       "  # 7 blank spaces

# Step 2: Extract all rows and columns
all_data = []

# Convert the sheet rows generator to a list to get the total row count
rows = list(sheet.iter_rows(values_only=True))

# Define the total width for the first two columns combined
total_width_before_last_column = 48  # 50 characters total for both columns combined

for row_index, row in enumerate(rows):
    # Handle the first two columns
    col1 = str(row[0]) if row[0] is not None else ""
    col2 = str(row[1]) if row[1] is not None else ""

    # Combine the first two columns with 1 space in between
    combined_columns = col1 + " " + col2

    # Ensure the combined columns are exactly 50 characters long
    combined_columns = combined_columns.ljust(total_width_before_last_column)[:total_width_before_last_column]

    # Handle the rest of the columns
    other_columns = ""
    if len(row) > 2:
        other_columns = " ".join([str(cell).strip() if cell is not None else "" for cell in row[2:-1]])

    last_column = str(row[-1]).strip()

    # Combine the formatted row
    formatted_row = combined_columns + " " + other_columns + " " + last_column

    # Add 7 blank spaces to all rows except the last one
    if row_index != len(rows) - 1:
        formatted_row += blank_spaces
    
    all_data.append(formatted_row)

# Step 3: Add the replacement string after all rows are processed
all_data.append(replacement_string)

# Step 4: Format the output
formatted_data = "\n".join(all_data)

# Step 5: Automatically generate the output file name
# Extract the base name of the Excel file and append '_output' to it
file_name, _ = os.path.splitext(os.path.basename(excel_file_path))
txt_file_path = os.path.join(os.path.dirname(excel_file_path), f"{file_name}_output.txt")

# Step 6: Save the data to the generated output file path
with open(txt_file_path, 'w') as file:
    file.write(formatted_data)

print(f"All data extracted and saved to {txt_file_path}")

# Step 7: Display column A in the console
print("\nColumn A from the processed data:")
for row in rows:
    col_a_value = str(row[0]) if row[0] is not None else ""
    print(col_a_value)
