from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import os

# Function to select the Excel file
def select_excel_file():
    """Prompts the user to select an Excel file and returns the file path."""
    Tk().withdraw()  # Hide the root tkinter window
    print("Opening file dialog to select an Excel file...")
    file_path = askopenfilename(
        title="Select Excel File",
        filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")]
    )
    if not file_path:
        print("No file selected. Exiting.")
        exit()  # Exit if no file is selected
    print(f"Selected file: {file_path}")
    return file_path

# Main processing function
def process_excel_and_save():
    # Step 1: Select the Excel file dynamically
    excel_file_path = select_excel_file()

    # Step 2: Load the workbook and process
    wb = load_workbook(excel_file_path)
    sheet = wb.active  # Or specify the sheet name: wb['Sheet1']

    replacement_string = '9999999999DEPARTMENT OF TRANSPORTAT000001079200266000010633854605905300000 00000'
    blank_spaces = "       "  # 7 blank spaces

    # Step 3: Extract and format data
    all_data = []
    rows = list(sheet.iter_rows(values_only=True))
    total_width_before_last_column = 48

    for row_index, row in enumerate(rows):
        col1 = str(row[0]) if row[0] is not None else ""
        col2 = str(row[1]) if row[1] is not None else ""
        combined_columns = (col1 + " " + col2).ljust(total_width_before_last_column)[:total_width_before_last_column]
        other_columns = " ".join([str(cell).strip() if cell is not None else "" for cell in row[2:-1]]) if len(row) > 2 else ""
        last_column = str(row[-1]).strip()
        formatted_row = combined_columns + " " + other_columns + " " + last_column

        if row_index != len(rows) - 1:
            formatted_row += blank_spaces

        all_data.append(formatted_row)

    # Handle the last row replacement
    if all_data:
        last_row_index = len(all_data) - 1
        last_row = rows[last_row_index]
        if any(cell is None for cell in last_row):
            all_data[last_row_index] = replacement_string

    # Save the output
    txt_file_path = os.path.splitext(excel_file_path)[0] + "_output.txt"
    with open(txt_file_path, 'w') as file:
        file.write("\n".join(all_data))

    print(f"Processed data saved to {txt_file_path}")

# Main entry point
if __name__ == "__main__":
    # Process Excel file
    process_excel_and_save()
